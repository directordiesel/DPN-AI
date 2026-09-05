from pathlib import Path

from openpyxl import load_workbook

from app.artifact_studio import ArtifactStudio


def test_artifact_studio_spreadsheet_adds_formula_chart_and_validation(tmp_path: Path):
    studio = ArtifactStudio(tmp_path)
    result = studio.create_spreadsheet(
        "finance",
        "Finance",
        [
            {
                "name": "Summary",
                "rows": [["Month", "Revenue"], ["Jan", 10], ["Feb", 15]],
                "formulas": {"B4": "SUM(B2:B3)"},
                "charts": [
                    {
                        "type": "bar",
                        "title": "Revenue",
                        "min_col": 2,
                        "max_col": 2,
                        "min_row": 1,
                        "max_row": 3,
                        "category_col": 1,
                        "anchor": "D2",
                    }
                ],
            }
        ],
    )

    assert result["ok"] is True
    assert result["validation"]["valid"] is True
    assert len(result["validation"]["sha256"]) == 64

    workbook = load_workbook(tmp_path / result["path"], data_only=False)
    ws = workbook["Summary"]
    assert ws["B4"].value == "=SUM(B2:B3)"
    assert len(ws._charts) == 1


def test_artifact_studio_document_returns_validation(tmp_path: Path):
    studio = ArtifactStudio(tmp_path)
    result = studio.create_document(
        "brief",
        "Brief",
        [{"heading": "Summary", "body": "Generated safely."}],
    )
    assert result["validation"]["valid"] is True
    assert result["validation"]["type"] == "docx"
