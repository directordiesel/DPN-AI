from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference

from app.artifact_validation import validate_artifact
from app.tools.documents import DocumentFactory


class ArtifactStudio:
    """Advanced artifact orchestration over the stable DocumentFactory API."""

    def __init__(self, workspace: Path):
        self.workspace = workspace.resolve()
        self.factory = DocumentFactory(self.workspace)

    def _finalize(self, result: dict[str, Any]) -> dict[str, Any]:
        if not result.get("ok") or not result.get("path"):
            return result
        validation = validate_artifact(self.workspace / str(result["path"]), self.workspace)
        return {**result, "validation": validation.to_dict()}

    def create_document(self, filename: str, title: str, sections: list[dict[str, Any]], author: str = "DPN AI") -> dict[str, Any]:
        return self._finalize(self.factory.create_docx(filename, title, sections, author=author))

    def create_pdf(self, filename: str, title: str, sections: list[dict[str, Any]]) -> dict[str, Any]:
        return self._finalize(self.factory.create_pdf(filename, title, sections))

    def create_presentation(self, filename: str, title: str, slides: list[dict[str, Any]]) -> dict[str, Any]:
        return self._finalize(self.factory.create_pptx(filename, title, slides))

    def create_spreadsheet(self, filename: str, title: str, sheets: list[dict[str, Any]]) -> dict[str, Any]:
        result = self.factory.create_xlsx(filename, title, sheets)
        if not result.get("ok") or not result.get("path"):
            return result

        target = (self.workspace / str(result["path"])).resolve()
        target.relative_to(self.workspace)
        workbook = load_workbook(target)

        for sheet_spec in sheets:
            name = str(sheet_spec.get("name", ""))[:31]
            if not name or name not in workbook.sheetnames:
                continue
            ws = workbook[name]

            formulas = sheet_spec.get("formulas", {})
            if isinstance(formulas, dict):
                for cell_ref, formula in formulas.items():
                    ref = str(cell_ref).strip().upper()
                    value = str(formula).strip()
                    if not ref or not value:
                        continue
                    ws[ref] = value if value.startswith("=") else f"={value}"

            charts = sheet_spec.get("charts", [])
            if not isinstance(charts, list):
                continue
            for chart_spec in charts[:8]:
                if not isinstance(chart_spec, dict):
                    continue
                chart_type = str(chart_spec.get("type", "bar")).lower()
                chart = LineChart() if chart_type == "line" else BarChart()
                chart.title = str(chart_spec.get("title", "Chart"))[:120]
                chart.y_axis.title = str(chart_spec.get("y_axis", ""))[:80]
                chart.x_axis.title = str(chart_spec.get("x_axis", ""))[:80]

                min_col = max(1, int(chart_spec.get("min_col", 2)))
                max_col = max(min_col, int(chart_spec.get("max_col", ws.max_column)))
                min_row = max(1, int(chart_spec.get("min_row", 1)))
                max_row = max(min_row + 1, int(chart_spec.get("max_row", ws.max_row)))
                category_col = max(1, int(chart_spec.get("category_col", 1)))
                data = Reference(ws, min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row)
                categories = Reference(ws, min_col=category_col, min_row=min_row + 1, max_row=max_row)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                chart.height = min(15, max(5, float(chart_spec.get("height", 7.5))))
                chart.width = min(25, max(8, float(chart_spec.get("width", 12))))
                ws.add_chart(chart, str(chart_spec.get("anchor", "H2")))

        workbook.save(target)
        return self._finalize(result)
